"""Trích xuất văn bản từ PDF (text/scan) và Excel. OCR bằng Tesseract (vie+eng)."""
from __future__ import annotations
from typing import Any, TypedDict
import io

import fitz  # PyMuPDF
from openpyxl import load_workbook


class PageText(TypedDict):
    page: int
    text: str


def classify_pdf(data: bytes) -> str:
    """Nếu tổng text trích được < ngưỡng -> coi là PDF scan."""
    doc = fitz.open(stream=data, filetype="pdf")
    total = sum(len(page.get_text().strip()) for page in doc)
    return "pdf_text" if total >= 20 else "pdf_scan"


def extract_pdf(data: bytes) -> list[PageText]:
    doc = fitz.open(stream=data, filetype="pdf")
    return [{"page": i + 1, "text": page.get_text()} for i, page in enumerate(doc)]


def ocr_pdf(data: bytes) -> list[PageText]:
    import pytesseract
    from PIL import Image

    doc = fitz.open(stream=data, filetype="pdf")
    out: list[PageText] = []
    for i, page in enumerate(doc):
        pix = page.get_pixmap(dpi=300)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        text = pytesseract.image_to_string(img, lang="vie+eng")
        out.append({"page": i + 1, "text": text})
    return out


class SheetData(TypedDict):
    sheet: str
    rows: list[list[Any]]


def parse_excel(data: bytes) -> list[SheetData]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    out: list[SheetData] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        out.append({"sheet": ws.title, "rows": rows})
    return out


def extract_document(data: bytes, file_kind: str) -> list[PageText]:
    if file_kind == "pdf_scan":
        return ocr_pdf(data)
    if file_kind == "pdf_text":
        return extract_pdf(data)
    if file_kind == "excel":
        pages: list[PageText] = []
        for i, sheet in enumerate(parse_excel(data)):
            text = "\n".join(
                "\t".join("" if c is None else str(c) for c in row)
                for row in sheet["rows"]
            )
            pages.append({"page": i + 1, "text": f"[{sheet['sheet']}]\n{text}"})
        return pages
    raise ValueError(f"file_kind không hỗ trợ: {file_kind}")
