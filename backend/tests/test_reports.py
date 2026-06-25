from decimal import Decimal
from pathlib import Path
from docx import Document
from openpyxl import load_workbook
from services import reports


def _sample():
    evals = {1: {
        "legality": [{"criteria_ten": "Đơn dự thầu", "result": "PASS", "score": 100,
                      "evidence": "Có chữ ký", "page_ref": [1], "note": "", "ai_model": "mock"}],
        "capacity": [], "technical": [],
        "financial": {"corrected_rows": [], "errors": [], "tong_gia": Decimal("200"),
                      "evaluated_price": Decimal("200")},
        "technical_score": 80.0, "passed_legality": True,
    }}
    ranking = [{"vendor_id": 1, "evaluated_price": Decimal("200"),
                "technical_score": 80.0, "rank": 1, "eligible": True}]
    return {"ma_so": "G-001", "ten": "Gói A"}, {1: "Công ty A"}, evals, ranking


def test_build_summary_docx(tmp_path):
    pkg, names, evals, ranking = _sample()
    out = reports.build_summary_docx(pkg, names, evals, ranking, tmp_path / "r.docx")
    assert out.exists()
    text = "\n".join(p.text for p in Document(str(out)).paragraphs)
    assert "Công ty A" in text and "G-001" in text


def test_build_summary_xlsx(tmp_path):
    pkg, names, evals, ranking = _sample()
    out = reports.build_summary_xlsx(names, evals, ranking, tmp_path / "r.xlsx")
    wb = load_workbook(out)
    assert "Xep hang" in wb.sheetnames
